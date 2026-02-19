"""Excel workbook report writer for QMeter analysis."""
from __future__ import annotations

import logging
import os
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_excel_report(
    outdir: str,
    filename: str,
    run_inventory: List[Dict],
    infra_table: List[Dict],
    score_summary: List[Dict],
    category_comparison: List[Dict],
    stability_table: List[Dict],
    outlier_table: List[Dict],
    findings: List[Dict],
    normalized_rows: List[Dict],
) -> str:
    """Write the complete Excel workbook report.

    Returns the path to the written file.
    """
    os.makedirs(outdir, exist_ok=True)
    filepath = os.path.join(outdir, filename)

    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 00_Run_Inventory
    _write_sheet(wb, "00_Run_Inventory", run_inventory)

    # 01_Infrastructure_Comparison
    _write_sheet(wb, "01_Infrastructure", infra_table)

    # 02_Score_Summary
    _write_sheet(wb, "02_Score_Summary", score_summary)

    # 03_Category_Summary
    _write_category_summary(wb, category_comparison)

    # Category-specific sheets
    _write_category_sheet(wb, "04_Memory_Test", category_comparison, "memory")
    _write_category_sheet(wb, "05_Algorithm_Test", category_comparison, "algorithm")
    _write_category_sheet(wb, "06_Quill_Speed_Test", category_comparison, "quill")
    _write_category_sheet(wb, "07_Propagator_Test", category_comparison, "propagator")
    _write_category_sheet(wb, "08_Database_Test", category_comparison, "database")

    # 09_Stability_Variance
    _write_sheet(wb, "09_Stability_Variance", stability_table)

    # 10_Outliers
    _write_sheet(wb, "10_Outliers", outlier_table)

    # 11_Findings_Text
    _write_findings_sheet(wb, findings)

    # 12_Data_Dictionary
    _write_data_dictionary(wb)

    wb.save(filepath)
    logger.info("Excel report written to %s", filepath)
    return filepath


def _write_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    rows: List[Dict],
    max_name_len: int = 31,
) -> None:
    """Write a generic sheet from a list of dicts."""
    name = sheet_name[:max_name_len]
    ws = wb.create_sheet(title=name)

    if not rows:
        ws.append(["No data available"])
        return

    headers = list(rows[0].keys())

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            value = row_data.get(header)
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

            # Format percentages
            if header in ("pct_change", "cv") and isinstance(value, (int, float)):
                cell.number_format = "0.00%"
                cell.value = value / 100.0 if abs(value) < 1000 else value

            # Color assessment cells
            if header == "assessment":
                if value == "regression":
                    cell.fill = BAD_FILL
                elif value == "improvement":
                    cell.fill = GOOD_FILL
                elif value == "unchanged":
                    cell.fill = NEUTRAL_FILL

    # Auto-width columns
    _auto_width(ws, headers)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions


def _write_category_summary(
    wb: openpyxl.Workbook,
    comparison_rows: List[Dict],
) -> None:
    """Write category summary sheet with all categories."""
    ws = wb.create_sheet(title="03_Category_Summary")

    if not comparison_rows:
        ws.append(["No comparison data available"])
        return

    headers = [
        "source_label", "is_baseline", "metric",
        "baseline_value", "current_value", "diff",
        "pct_change", "assessment", "directionality",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(comparison_rows, 2):
        for col, header in enumerate(headers, 1):
            value = row_data.get(header)
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

            if header == "pct_change" and isinstance(value, (int, float)):
                cell.number_format = "0.00"
                cell.value = value

            if header == "assessment":
                if value == "regression":
                    cell.fill = BAD_FILL
                elif value == "improvement":
                    cell.fill = GOOD_FILL

    _auto_width(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_category_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    comparison_rows: List[Dict],
    category_filter: str,
) -> None:
    """Write a category-specific comparison sheet."""
    filtered = [r for r in comparison_rows if category_filter in r.get("metric", "")]
    _write_sheet(wb, sheet_name, filtered)


def _write_findings_sheet(
    wb: openpyxl.Workbook,
    findings: List[Dict],
) -> None:
    """Write findings as a formatted sheet."""
    ws = wb.create_sheet(title="11_Findings_Text")

    if not findings:
        ws.append(["No significant findings detected."])
        return

    headers = ["#", "severity", "category", "title", "detail", "recommendation"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, finding in enumerate(findings, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=finding.get("severity", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=finding.get("category", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=finding.get("title", "")).border = THIN_BORDER

        detail_cell = ws.cell(row=row_idx, column=5, value=finding.get("detail", ""))
        detail_cell.border = THIN_BORDER
        detail_cell.alignment = Alignment(wrap_text=True)

        rec_cell = ws.cell(row=row_idx, column=6, value=finding.get("recommendation", ""))
        rec_cell.border = THIN_BORDER
        rec_cell.alignment = Alignment(wrap_text=True)

        # Color severity
        severity = finding.get("severity", "")
        sev_cell = ws.cell(row=row_idx, column=2)
        if severity in ("critical", "high"):
            sev_cell.fill = BAD_FILL
        elif severity == "medium":
            sev_cell.fill = NEUTRAL_FILL
        elif severity in ("low", "info"):
            sev_cell.fill = GOOD_FILL

    _auto_width(ws, headers)
    ws.freeze_panes = "A2"


def _write_data_dictionary(wb: openpyxl.Workbook) -> None:
    """Write data dictionary sheet."""
    ws = wb.create_sheet(title="12_Data_Dictionary")

    entries = [
        ("score1", "Overall QMeter score with 1 dataset (single-threaded)", "Lower is better", "milliseconds"),
        ("score4", "Overall QMeter score with 4 datasets", "Lower is better", "milliseconds"),
        ("score16", "Overall QMeter score with 16 datasets", "Lower is better", "milliseconds"),
        ("avg_runtime", "Average runtime across tasks in a category/group", "Lower is better", "milliseconds"),
        ("group_score", "Group score = AvgRuntime + StdDevRuntime (penalizes variance)", "Lower is better", "milliseconds"),
        ("std_dev_runtime", "Standard deviation of task runtimes within a group", "Lower is better", "milliseconds"),
        ("Memory test", "Category 1: Tests memory performance using WPBL data creation and propagation", "Lower runtime is better", "milliseconds"),
        ("Algorithm test", "Category 2: Tests POA, CP, and MP optimization algorithms", "Lower runtime is better", "milliseconds"),
        ("Quill speed test", "Category 3: Tests Quill logic speed (Hanoi Tower test)", "Lower runtime is better", "milliseconds"),
        ("Propagator test", "Category 4: Tests propagator behavior in medium transactions", "Lower runtime is better", "milliseconds"),
        ("Database test", "Category 5: Tests data storage performance (full/partial/memory)", "Lower runtime is better", "milliseconds"),
        ("NrOfThreads", "Number of parallel threads/datasets used in a test group", "N/A", "count"),
        ("CV (Coefficient of Variation)", "Ratio of std dev to mean, expressed as percentage", "Lower is more stable", "percent"),
        ("p90", "90th percentile value", "Context-dependent", "same as metric"),
        ("p95", "95th percentile value", "Context-dependent", "same as metric"),
        ("ci_95_lower/upper", "95% confidence interval bounds for the mean", "Narrower is more precise", "same as metric"),
        ("pct_change", "Percentage change vs baseline", "Positive = worse for durations", "percent"),
    ]

    headers = ["Metric", "Description", "Directionality", "Unit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row_idx, (metric, desc, direction, unit) in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=direction).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=unit).border = THIN_BORDER

    _auto_width(ws, headers)
    ws.freeze_panes = "A2"


def _auto_width(ws, headers: List[str], min_width: int = 10, max_width: int = 50) -> None:
    """Auto-adjust column widths."""
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))

        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width
